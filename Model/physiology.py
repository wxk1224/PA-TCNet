import csv
import json
from pathlib import Path
from typing import Dict, List

import numpy as np


ROOT = Path(__file__).resolve().parents[2]
REQUIRED_ROI_GROUPS = ("left", "right", "midline")


def read_channel_info(info_path):
    info_path = Path(info_path)
    with open(info_path, "r", encoding="utf-8") as handle:
        channel_info = json.load(handle)

    channel_names = channel_info.get("channel_names")
    roi_groups = channel_info.get("roi_groups")
    if not isinstance(channel_names, list) or not channel_names:
        raise ValueError(f"channel_names must be a non-empty list in {info_path}")
    if not isinstance(roi_groups, dict):
        raise ValueError(f"roi_groups must be an object in {info_path}")

    missing_groups = [group_name for group_name in REQUIRED_ROI_GROUPS if group_name not in roi_groups]
    if missing_groups:
        raise ValueError(f"Missing ROI groups in {info_path}: {missing_groups}")

    return {
        "channel_names": [str(name) for name in channel_names],
        "roi_groups": {
            group_name: [str(name) for name in roi_groups[group_name]]
            for group_name in REQUIRED_ROI_GROUPS
        },
    }


def _resolve_indices(channel_names, roi_names):
    channel_to_index = {name: idx for idx, name in enumerate(channel_names)}
    missing = [name for name in roi_names if name not in channel_to_index]
    if missing:
        raise ValueError(f"Missing ROI channels in channel_names: {missing}")
    return [channel_to_index[name] for name in roi_names]


def get_motor_roi_indices(channel_info):
    channel_names = channel_info["channel_names"]
    roi_groups = channel_info["roi_groups"]
    return {
        "left": _resolve_indices(channel_names, roi_groups["left"]),
        "right": _resolve_indices(channel_names, roi_groups["right"]),
        "midline": _resolve_indices(channel_names, roi_groups["midline"]),
    }


def l2_normalize(vectors, axis=1, eps=1e-8):
    norms = np.linalg.norm(vectors, axis=axis, keepdims=True)
    return vectors / np.clip(norms, eps, None)


def compute_roi_features(x, channel_info):
    x = np.asarray(x, dtype=np.float32)
    if x.ndim != 3:
        raise ValueError(f"Expected input shape [N, C, T], got {x.shape}")

    roi_indices = get_motor_roi_indices(channel_info)

    left_motor = x[:, roi_indices["left"], :].mean(axis=1)
    right_motor = x[:, roi_indices["right"], :].mean(axis=1)
    midline = x[:, roi_indices["midline"], :].mean(axis=1)
    asymmetry = np.abs(left_motor - right_motor)

    roi_vector = np.concatenate([left_motor, right_motor, asymmetry, midline], axis=1)
    roi_vector = l2_normalize(roi_vector, axis=1)

    return {
        "left_motor": left_motor,
        "right_motor": right_motor,
        "midline": midline,
        "asymmetry": asymmetry,
        "roi_vector": roi_vector.astype(np.float32),
    }


def build_class_templates(roi_vectors, labels, num_classes=2, floor=0.70):
    roi_vectors = np.asarray(roi_vectors, dtype=np.float32)
    labels = np.asarray(labels, dtype=np.int64)

    templates = []
    thresholds = []

    for class_id in range(num_classes):
        class_vectors = roi_vectors[labels == class_id]
        if len(class_vectors) == 0:
            raise ValueError(f"No source samples found for class {class_id}.")

        template = class_vectors.mean(axis=0, keepdims=True)
        template = l2_normalize(template, axis=1)[0]

        similarities = class_vectors @ template
        threshold = max(float(similarities.mean() - similarities.std()), float(floor))

        templates.append(template.astype(np.float32))
        thresholds.append(threshold)

    return np.stack(templates, axis=0), np.asarray(thresholds, dtype=np.float32)


def _normalize_side(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"l", "left", "left hand", "left-hand"}:
        return "left"
    if text in {"r", "right", "right hand", "right-hand"}:
        return "right"
    return None


def _normalize_float(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _header_key(value):
    return str(value).strip().lower().replace(" ", "")


def _load_xw_metadata(xlsx_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read XW metadata.") from exc

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"XW metadata file is empty: {xlsx_path}")

    headers = {_header_key(name): idx for idx, name in enumerate(rows[0]) if name is not None}
    idx_col = headers.get("idx")
    side_col = headers.get("side")
    dominant_col = headers.get("dominant")
    mrs_col = headers.get("mrs")

    if idx_col is None or side_col is None:
        raise ValueError(f"XW metadata file is missing required columns: {xlsx_path}")

    metadata = {}
    for row in rows[1:]:
        if row[idx_col] is None:
            continue
        subject_id = int(row[idx_col])
        paralysis_side = _normalize_side(row[side_col])
        dominant_hand = _normalize_side(row[dominant_col]) if dominant_col is not None else None
        mrs = _normalize_float(row[mrs_col]) if mrs_col is not None else None
        metadata[subject_id] = {
            "subject_id": subject_id,
            "paralysis_side": paralysis_side,
            "dominant_hand": dominant_hand,
            "mrs": mrs,
            "need_flip": paralysis_side == "left",
        }
    return metadata


def _load_ty_metadata(csv_path):
    metadata = {}
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            subject_id_text = str(row.get("subject_id", "")).strip()
            if not subject_id_text:
                continue
            subject_id = int(subject_id_text)
            paralysis_side = _normalize_side(row.get("paralysis_side"))
            dominant_hand = _normalize_side(row.get("dominant_hand"))
            mrs = _normalize_float(row.get("mrs"))
            metadata[subject_id] = {
                "subject_id": subject_id,
                "paralysis_side": paralysis_side,
                "dominant_hand": dominant_hand,
                "mrs": mrs,
                "need_flip": paralysis_side == "left",
            }
    return metadata


def load_subject_metadata(dataset_name, xw_metadata_path=None, ty_metadata_path=None):
    if dataset_name.startswith("XW"):
        xw_metadata_path = Path(xw_metadata_path or ROOT / "assets" / "XwSubInfo.xlsx")
        if not xw_metadata_path.exists():
            raise FileNotFoundError(f"XW metadata file not found: {xw_metadata_path}")
        return _load_xw_metadata(xw_metadata_path)

    if dataset_name.startswith("TY"):
        ty_metadata_path = Path(ty_metadata_path or ROOT / "assets" / "TySubInfo.csv")
        if not ty_metadata_path.exists():
            raise FileNotFoundError(f"TY metadata file not found: {ty_metadata_path}")
        return _load_ty_metadata(ty_metadata_path)

    raise ValueError(f"Unsupported dataset for metadata loading: {dataset_name}")


def validate_subject_metadata(subject_metadata: Dict[int, Dict], subject_ids: List[int], dataset_name):
    missing_subjects = [sid for sid in subject_ids if sid not in subject_metadata]
    if missing_subjects:
        raise ValueError(f"{dataset_name} metadata is missing subjects: {missing_subjects}")

    invalid_sides = [
        sid
        for sid in subject_ids
        if subject_metadata[sid].get("paralysis_side") not in {"left", "right"}
    ]
    if invalid_sides:
        raise ValueError(
            f"{dataset_name} metadata has missing or invalid paralysis_side for subjects: {invalid_sides}"
        )


def _remap_labels_to_affected_unaffected(y, paralysis_side):
    y = np.asarray(y, dtype=np.int64)
    unique_labels = set(np.unique(y).tolist())
    if not unique_labels.issubset({0, 1}):
        raise ValueError(f"Expected binary labels 0/1, got {sorted(unique_labels)}")

    if paralysis_side == "left":
        affected_mask = y == 0
        unaffected_mask = y == 1
    elif paralysis_side == "right":
        affected_mask = y == 1
        unaffected_mask = y == 0
    else:
        raise ValueError(f"Invalid paralysis_side: {paralysis_side}")

    remapped = np.full_like(y, fill_value=-1)
    remapped[affected_mask] = 0
    remapped[unaffected_mask] = 1
    if np.any(remapped < 0):
        raise ValueError("Failed to remap labels to affected/unaffected.")
    return remapped


def canonicalize_subject_trials(x, y, subject_id, subject_metadata, flip_indices=None):
    if subject_id not in subject_metadata:
        raise ValueError(f"Subject {subject_id} is missing from metadata.")

    meta = dict(subject_metadata[subject_id])
    paralysis_side = meta.get("paralysis_side")
    if paralysis_side not in {"left", "right"}:
        raise ValueError(f"Subject {subject_id} has invalid paralysis_side: {paralysis_side}")

    x = np.asarray(x, dtype=np.float32)
    y = np.asarray(y, dtype=np.int64)
    if x.ndim != 3:
        raise ValueError(f"Expected subject trials with shape [N, C, T], got {x.shape}")

    canonical_x = x.copy()
    if meta.get("need_flip", False):
        if flip_indices is None:
            raise ValueError("flip_indices must be provided when channel flipping is required.")
        flip_indices = np.asarray(flip_indices, dtype=np.int64)
        canonical_x = canonical_x[:, flip_indices, :]

    canonical_y = _remap_labels_to_affected_unaffected(y, paralysis_side)
    meta["label_space"] = {"0": "affected", "1": "unaffected"}
    return canonical_x, canonical_y, meta
